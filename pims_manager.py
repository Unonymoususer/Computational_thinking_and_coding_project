import os
import webbrowser
from openpyxl import Workbook
from openpyxl import load_workbook
import googlemaps
import xlsxwriter
import xlrd
import xlwt

def 시작():
    선택 = 0

    while 선택 != 7:
        선택 = 메뉴선택()

        if 선택 == 1:
            등록()
        elif 선택 == 2:
            전체검색()
        elif 선택 == 3:
            개인검색()
        elif 선택 == 4:
            수정()
        elif 선택 == 5:
            삭제()
        elif 선택 == 6:
            전체삭제()


def 메뉴선택():
    print()
    print('학생 정보 관리 메뉴')
    print('-------------------')
    print('1. 등록')
    print('2. 전체검색')
    print('3. 개인검색')
    print('4. 수정')
    print('5. 삭제(개인)')
    print('6. 삭제(전체)')
    print('7. 종료')
    print()

    선택 = int(input('선택 >> '))

    while 선택 < 1 or 선택 > 7:
        선택 = int(input('메뉴 번호를 확인해주세요: '))

    return 선택


def 등록():
    print()
    print('메모장과 엑셀파일에 학생의 정보를 등록합니다.')
    학번 = input('학번: ')
    이름 = input('이름: ')
    학과 = input('학과: ')
    주소 = input('주소: ')

    등록_메모장(학번,이름,학과,주소)
    등록_엑셀(학번,이름,학과,주소)

    print('학생정보의 등록이 완료되었습니다.')


def 전체검색():
    print()    
    읽기모드파일 = open('학생정보.txt', 'r')

    학번 = 읽기모드파일.readline()
    while 학번 != '':
        이름 = 읽기모드파일.readline()
        학과 = 읽기모드파일.readline()
        주소 = 읽기모드파일.readline()

        학번 = 학번.rstrip('\n')
        이름 = 이름.rstrip('\n')
        학과 = 학과.rstrip('\n')
        주소 = 주소.rstrip('\n')

        print(학번 + '\t' + 이름 + '\t' + 학과 + '\t' + 주소)
        학번 = 읽기모드파일.readline()

    읽기모드파일.close()


def 개인검색():
    print()
    검색선택 = 0
    학번리스트 = []
    이름리스트 = []
    학과리스트 = []
    주소리스트 = []
    검색선택 = 개인검색_메뉴(검색선택)

    if 검색선택 == 1:
        입력데이터 = input('학번을 입력해주세요: ')

    elif 검색선택 == 2:
        입력데이터 = input('이름을 입력해주세요: ')

    elif 검색선택 == 3:
        입력데이터 = input('학과명을 입력해주세요: ')

    elif 검색선택 == 4:
        입력데이터 = input('주소를 입력해주세요: ')

    
    읽기모드파일 = open('학생정보.txt', 'r')

    학번 = 읽기모드파일.readline()
    발견 = '발견못함'
    
    while 학번 != '':
        이름 = 읽기모드파일.readline()
        학과 = 읽기모드파일.readline()
        주소 = 읽기모드파일.readline()

        학번 = 학번.rstrip('\n')
        이름 = 이름.rstrip('\n')
        학과 = 학과.rstrip('\n')
        주소 = 주소.rstrip('\n')

        if 검색선택 == 1:
            if 입력데이터 in 학번:
                print(학번 + '\t' + 이름 + '\t' + 학과 + '\t' + 주소)
                학번목록 = [학번]
                이름목록 = [이름]
                학과목록 = [학과]
                주소목록 = [주소]
                학번리스트 = 학번리스트 + 학번목록
                이름리스트 = 이름리스트 + 이름목록
                학과리스트 = 학과리스트 + 학과목록
                주소리스트 = 주소리스트 + 주소목록
                발견 = '발견함'

        elif 검색선택 == 2:
            if 입력데이터 in 이름:
                print(학번 + '\t' + 이름 + '\t' + 학과 + '\t' + 주소)
                학번목록 = [학번]
                이름목록 = [이름]
                학과목록 = [학과]
                주소목록 = [주소]
                학번리스트 = 학번리스트 + 학번목록
                이름리스트 = 이름리스트 + 이름목록
                학과리스트 = 학과리스트 + 학과목록
                주소리스트 = 주소리스트 + 주소목록
                발견 = '발견함'

        elif 검색선택 == 3:
            if 입력데이터 in 학과:
                print(학번 + '\t' + 이름 + '\t' + 학과 + '\t' + 주소)
                학번목록 = [학번]
                이름목록 = [이름]
                학과목록 = [학과]
                주소목록 = [주소]
                학번리스트 = 학번리스트 + 학번목록
                이름리스트 = 이름리스트 + 이름목록
                학과리스트 = 학과리스트 + 학과목록
                주소리스트 = 주소리스트 + 주소목록
                발견 = '발견함'

        elif 검색선택 == 4:
            if 입력데이터 in 주소:
                print(학번 + '\t' + 이름 + '\t' + 학과 + '\t' + 주소)
                학번목록 = [학번]
                이름목록 = [이름]
                학과목록 = [학과]
                주소목록 = [주소]
                학번리스트 = 학번리스트 + 학번목록
                이름리스트 = 이름리스트 + 이름목록
                학과리스트 = 학과리스트 + 학과목록
                주소리스트 = 주소리스트 + 주소목록
                발견 = '발견함'


        학번 = 읽기모드파일.readline()

    if 발견 == '발견함':
        #개인검색_추가설정()
        개인검색_추가설정(학번리스트,이름리스트,학과리스트,주소리스트)

    elif 발견 == '발견못함':
        print('해당 검색조건에 맞는 정보가 없습니다.')

    읽기모드파일.close()


def 수정():
    print()    
    입력받은_학번 = input('학번을 입력해주세요: ')
    입력받은_주소 = input(입력받은_학번 + '의 새로운 주소를 입력해주세요: ')

    읽기모드파일 = open('학생정보.txt', 'r')
    임시파일 = open('임시파일.txt', 'w')
    
    학번 = 읽기모드파일.readline()
    발견 = '발견못함'
    while 학번 != '':
        이름 = 읽기모드파일.readline()
        학과 = 읽기모드파일.readline()
        주소 = 읽기모드파일.readline()

        학번 = 학번.rstrip('\n')
        이름 = 이름.rstrip('\n')
        학과 = 학과.rstrip('\n')
        주소 = 주소.rstrip('\n')

        if 입력받은_학번 == 학번:
            임시파일.write(학번 + '\n')
            임시파일.write(이름 + '\n')
            임시파일.write(학과 + '\n')
            임시파일.write(입력받은_주소 + '\n')
            발견 = '발견함'
        else:
            임시파일.write(학번 + '\n')
            임시파일.write(이름 + '\n')
            임시파일.write(학과 + '\n')
            임시파일.write(주소 + '\n')

        학번 = 읽기모드파일.readline()

    읽기모드파일.close()
    임시파일.close()

    os.remove('학생정보.txt')
    os.rename('임시파일.txt', '학생정보.txt')

    if 발견 == '발견함':
        print('학생 정보가 업데이트되었습니다.')
    else:
        print('학번에 대한 정보가 없어 업데이트를 하지 못하였습니다.')      


def 삭제():
    print()
    입력받은_학번 = input('학번을 입력해주세요: ')

    읽기모드파일 = open('학생정보.txt', 'r')
    임시파일 = open('임시파일.txt', 'w')
    
    학번 = 읽기모드파일.readline()
    발견 = '발견못함'
    while 학번 != '':
        이름 = 읽기모드파일.readline()
        학과 = 읽기모드파일.readline()
        주소 = 읽기모드파일.readline()

        학번 = 학번.rstrip('\n')
        이름 = 이름.rstrip('\n')        
        학과 = 학과.rstrip('\n')
        주소 = 주소.rstrip('\n')

        if 입력받은_학번 == 학번:
            발견 = '발견함'
        else:
            임시파일.write(학번 + '\n')
            임시파일.write(이름 + '\n')
            임시파일.write(학과 + '\n')
            임시파일.write(주소 + '\n')

        학번 = 읽기모드파일.readline()

    읽기모드파일.close()
    임시파일.close()

    os.remove('학생정보.txt')
    os.rename('임시파일.txt', '학생정보.txt')

    if 발견 == '발견함':
        print('회원 정보가 삭제되었습니다.')
    else:
        print('학번에 대한 정보가 없어 삭제하지 못하였습니다.')


def 전체삭제():
    print()
    print('삭제할 파일을 선택해주세요.')
    print('1. 학생정보(메모장)')
    print('2. 학생정보(엑셀)')
    print('3. 개인검색(메모장)')
    print('4. 개인검색(엑셀)')
    print('5. 전체삭제')

    응답 = int(input('선택 >> '))

    while 응답 < 1 or 응답 > 5:
        응답 = int(input('메뉴 번호를 확인해주세요: '))
    
    if 응답 == '1':
        os.remove('학생정보.txt')
        쓰기모드파일 = open('학생정보.txt', 'w')
        쓰기모드파일.close()

    elif 응답 == '2':
        os.remove('학생정보.xlsx')
        쓰기모드파일 = open('학생정보.xlsx', 'w')
        쓰기모드파일.close()

    elif 응답 == '3':
        os.remove('개인검색.txt')
        쓰기모드파일 = open('개인검색.txt', 'w')
        쓰기모드파일.close()

    elif 응답 == '4':
        os.remove('개인검색.xlsx')
        쓰기모드파일 = open('개인검색.xlsx', 'w')
        쓰기모드파일.close()

    elif 응답 == '5':
        os.remove('학생정보.txt')
        쓰기모드파일 = open('학생정보.txt', 'w')
        쓰기모드파일.close()
        os.remove('학생정보.xlsx')
        쓰기모드파일 = open('학생정보.xlsx', 'w')
        쓰기모드파일.close()
        os.remove('개인검색.txt')
        쓰기모드파일 = open('개인검색.txt', 'w')
        쓰기모드파일.close()
        os.remove('개인검색.xlsx')
        쓰기모드파일 = open('개인검색.xlsx', 'w')
        쓰기모드파일.close()  


def 등록_메모장(학번,이름,학과,주소):
    쓰기모드파일_메모장 = open('학생정보.txt', 'a')

    쓰기모드파일_메모장.write(학번 + '\n')
    쓰기모드파일_메모장.write(이름 + '\n')
    쓰기모드파일_메모장.write(학과 + '\n')
    쓰기모드파일_메모장.write(주소 + '\n')
    쓰기모드파일_메모장.close()


def 등록_엑셀(학번,이름,학과,주소):
    
    워크북_읽기 = xlrd.open_workbook('학생정보.xlsx')
    워크시트 = 워크북_읽기.sheet_by_index(0)
    학생수 = 워크시트.nrows
    
    워크북 = load_workbook('학생정보.xlsx')
    워크시트 = 워크북.active
    워크시트.title = '학생정보'
    워크시트.cell(row = 학생수+1, column = 1).value = int(학번)
    워크시트.cell(row = 학생수+1, column = 2).value = 이름
    워크시트.cell(row = 학생수+1, column = 3).value = 학과
    워크시트.cell(row = 학생수+1, column = 4).value = 주소

    워크북.save('학생정보.xlsx')

    

def 개인검색_메뉴(검색선택):
    print('검색 방법을 골라주세요.')
    print('1. 학번으로 검색')
    print('2. 이름으로 검색')
    print('3. 학과명으로 검색')
    print('4. 주소지로 검색')

    검색선택 = int(input('선택 >> '))

    while 검색선택 < 1 or 검색선택 > 4:
        검색선택 = int(input('메뉴 번호를 확인해주세요: '))

    if 검색선택 == 1:
        print()
        print('학번으로 검색합니다.')

    elif 검색선택 == 2:
        print()
        print('이름으로 검색합니다.')

    elif 검색선택 == 3:
        print()
        print('학과명으로 검색합니다.')

    elif 검색선택 == 4:
        print()
        print('주소지로 검색합니다.')

    return 검색선택

        


def 개인검색_추가설정(학번리스트,이름리스트,학과리스트,주소리스트):
    print()
    print('추가작업을 수행할 수 있습니다.')
    print('------------------------------')
    print('1. 추가작업을 하지않고 넘어감')
    print('2. 지도로 주소검색')
    print('3. 검색한 인물을 별도로 저장(메모장)')
    print('4. 검색한 인물을 별도로 저장(엑셀)')

    추가선택 = int(input('선택 >> '))

    while 추가선택 < 1 or 추가선택 > 4:
        추가선택 = int(input('메뉴 번호를 확인해주세요: '))         

    if 추가선택 == 1:
        print()
        print('추가작업을 수행하지 않고 넘어갑니다.')

    elif 추가선택 == 2:
        print()
        print('지도로 주소를 검색합니다.')
        개인검색_추가설정_인터넷검색(학번리스트,이름리스트,학과리스트,주소리스트)

    elif 추가선택 == 3:
        print()
        print('검색한 인물을 별도로 저장합니다(메모장)')
        개인검색_추가설정_별도저장_메모장(학번리스트,이름리스트,학과리스트,주소리스트)

    elif 추가선택 == 4:
        print()
        print('검색한 인물을 별도로 저장합니다(엑셀)')
        개인검색_추가설정_별도저장_엑셀(학번리스트,이름리스트,학과리스트,주소리스트)


  
def 개인검색_추가설정_인터넷검색(학번리스트,이름리스트,학과리스트,주소리스트):
    print('임시데이터')


def 개인검색_추가설정_별도저장_메모장(학번리스트,이름리스트,학과리스트,주소리스트):
    os.remove('개인검색.txt')
    쓰기모드파일_메모장 = open('개인검색.txt', 'w')
    학생순번 = 0
    학생수 = len(학번리스트)

    while 학생순번 != 학생수:
        
        for 순번 in 학번리스트:
            쓰기모드파일_메모장.write(순번 + '\n')
            del 학번리스트[0]
            break

        for 순번 in 이름리스트:
            쓰기모드파일_메모장.write(순번 + '\n')
            del 이름리스트[0]
            break

        for 순번 in 학과리스트:
            쓰기모드파일_메모장.write(순번 + '\n')
            del 학과리스트[0]
            break

        for 순번 in 주소리스트:
            쓰기모드파일_메모장.write(순번 + '\n')
            del 주소리스트[0]
            break


        학생순번 = 학생순번 + 1
            
    
    쓰기모드파일_메모장.close()


def 개인검색_추가설정_별도저장_엑셀(학번리스트,이름리스트,학과리스트,주소리스트):
    os.remove('개인검색.xlsx')
    엑셀파일생성 = Workbook()
    워크시트생성 = 엑셀파일생성.active
    워크시트생성.title = '학생정보'

    학생순번 = 0
    
    for 순번 in 학번리스트:
        워크시트생성.cell(row = 학생순번+1, column = 1).value = int(순번)
        학생순번 = 학생순번 + 1
    학생순번 = 0
    
    for 순번 in 이름리스트:
        워크시트생성.cell(row = 학생순번+1, column = 2).value = 순번
        학생순번 = 학생순번 + 1
    학생순번 = 0
    
    for 순번 in 학과리스트:
        워크시트생성.cell(row = 학생순번+1, column = 3).value = 순번
        학생순번 = 학생순번 + 1
    학생순번 = 0
    
    for 순번 in 주소리스트:
        워크시트생성.cell(row = 학생순번+1, column = 4).value = 순번
        학생순번 = 학생순번 + 1

    엑셀파일생성.save('개인검색.xlsx')

    

시작()
